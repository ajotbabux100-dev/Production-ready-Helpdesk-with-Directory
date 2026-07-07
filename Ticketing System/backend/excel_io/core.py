"""Shared helpers for the "master upload" Excel template/import feature used
by Directory, Portals and Users. Kept generic on purpose - each app supplies
its own headers/notes and its own row-by-row create-or-update logic; this
module only handles the .xlsx plumbing (building a template workbook,
reading an uploaded workbook back into plain rows).
"""
import base64
import io

import openpyxl
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
from django.http import HttpResponse

DATA_SHEET_NAME = 'Data'


class BadUpload(Exception):
    """Raised when the uploaded file isn't a readable .xlsx workbook, so
    callers can turn it into a clean 400 instead of a 500."""

HEADER_FONT = Font(bold=True, color='FFFFFF')
HEADER_FILL = PatternFill(start_color='1E3A5F', end_color='1E3A5F', fill_type='solid')

# Leading characters Excel/LibreOffice/Sheets treat as "this cell is a
# formula" when the file is opened - a value like "=cmd|'/c calc'!A1" typed
# into a ticket title, directory entry, or imported spreadsheet round-trips
# untouched into every report/export we generate and would execute on open
# (CSV/Excel injection). Neutralized by prefixing with a leading apostrophe,
# which Excel renders as plain text and openpyxl writes as a literal value
# rather than a formula.
_FORMULA_PREFIXES = ('=', '+', '-', '@')


def _sanitize_cell(value):
    if isinstance(value, str) and value[:1] in _FORMULA_PREFIXES:
        return "'" + value
    return value


def _sanitize_row(row):
    return [_sanitize_cell(v) for v in row]


def _style_header_row(ws, headers):
    ws.append(list(headers))
    for cell in ws[1]:
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
    for col_idx, header in enumerate(headers, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = max(16, len(str(header)) + 4)
    ws.freeze_panes = 'A2'


def _as_response(wb, filename):
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response


def build_template(filename, headers, notes=None, sample_rows=None):
    """Builds an .xlsx template with a header row (and optional sample rows)
    on a "Data" sheet, plus an optional "Instructions" sheet, and returns it
    as a downloadable HttpResponse."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = DATA_SHEET_NAME
    _style_header_row(ws, headers)

    for row in (sample_rows or []):
        ws.append(_sanitize_row(row))

    if notes:
        notes_ws = wb.create_sheet('Instructions')
        notes_ws['A1'] = 'Instructions'
        notes_ws['A1'].font = Font(bold=True, size=13)
        for i, note in enumerate(notes, start=3):
            notes_ws.cell(row=i, column=1, value=note)
        notes_ws.column_dimensions['A'].width = 100

    return _as_response(wb, filename)


def build_workbook(filename, sheets):
    """Builds a multi-sheet .xlsx for data export (as opposed to
    build_template's single re-uploadable "Data" sheet). `sheets` is a list
    of (sheet_name, headers, rows) tuples - one sheet per report/table."""
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    for sheet_name, headers, rows in sheets:
        ws = wb.create_sheet(sheet_name[:31])  # Excel sheet name limit
        _style_header_row(ws, headers)
        for row in rows:
            ws.append(_sanitize_row(row))
    return _as_response(wb, filename)


def build_import_report_base64(headers, created_rows, skipped_rows, error_rows):
    """Builds a full import report with one sheet each for Created, Skipped
    (Already Exists) and Errors (only the non-empty ones are included) - so
    "which entries were created vs. skipped" is a reviewable file, not just
    two numbers in a toast. Imports are create-only: a row matching an
    existing record is skipped, never used to modify it.

    created_rows/skipped_rows: list of {'row': int, 'data': sequence} dicts.
    error_rows: list of {'row': int, 'error': str, 'data': sequence} dicts.
    All 'data' sequences align with `headers`."""
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    if created_rows:
        ws = wb.create_sheet('Created')
        _style_header_row(ws, ['Row', *headers])
        for r in created_rows:
            ws.append(_sanitize_row([r['row'], *(r.get('data') or [])]))

    if skipped_rows:
        ws = wb.create_sheet('Skipped (Already Exists)')
        _style_header_row(ws, ['Row', *headers])
        for r in skipped_rows:
            ws.append(_sanitize_row([r['row'], *(r.get('data') or [])]))

    if error_rows:
        ws = wb.create_sheet('Errors')
        _style_header_row(ws, ['Row', 'Error', *headers])
        for err in error_rows:
            ws.append(_sanitize_row([err['row'], err['error'], *(err.get('data') or [])]))

    if not wb.sheetnames:
        wb.create_sheet('Report')

    buf = io.BytesIO()
    wb.save(buf)
    return base64.b64encode(buf.getvalue()).decode('ascii')


def read_upload(file_obj):
    """Reads an uploaded .xlsx (the "Data" sheet if present, otherwise the
    first sheet). Returns (headers, rows) where headers is a list of
    stripped strings and rows is a list of raw cell-value tuples, skipping
    any fully-blank rows. Raises BadUpload if the file isn't a valid .xlsx
    (e.g. wrong file type, corrupted upload) - callers should catch this and
    return a 400 rather than letting it surface as an unhandled 500."""
    name = getattr(file_obj, 'name', '') or ''
    if name and not name.lower().endswith('.xlsx'):
        raise BadUpload('File must be an .xlsx spreadsheet.')
    try:
        # Deliberately NOT read_only=True: in that mode openpyxl trusts the
        # worksheet's <dimension> XML metadata to know where the data ends,
        # which can be stale (e.g. after editing/re-saving in Google Sheets,
        # LibreOffice, or Excel itself) - that silently truncates iter_rows()
        # to far fewer rows than the file actually has, with no error at all.
        # Loading fully avoids trusting that metadata. These are admin bulk-
        # upload files (hundreds to a few thousand rows at most), not a case
        # where read_only's memory savings matter more than correctness.
        wb = openpyxl.load_workbook(file_obj, data_only=True, read_only=False)
    except Exception as e:
        raise BadUpload(f'Could not read this file as an Excel spreadsheet: {e}')
    ws = wb[DATA_SHEET_NAME] if DATA_SHEET_NAME in wb.sheetnames else wb.worksheets[0]

    rows_iter = ws.iter_rows(values_only=True)
    try:
        header_row = next(rows_iter)
    except StopIteration:
        return [], []
    headers = [str(h).strip() if h is not None else '' for h in header_row]

    data_rows = []
    for row in rows_iter:
        if all(cell is None or str(cell).strip() == '' for cell in row):
            continue
        data_rows.append(row)
    return headers, data_rows


def row_cell(headers, row, name):
    """Looks up a named column's value in a raw row tuple by header name."""
    if name not in headers:
        return ''
    idx = headers.index(name)
    if idx >= len(row) or row[idx] is None:
        return ''
    return str(row[idx]).strip()
